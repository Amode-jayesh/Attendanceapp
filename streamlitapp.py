import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

# Function to calculate attendance percentage
def calculate_attendance_percentage(df):
    df['Attendance Percentage'] = (df['Classes Attended'] / df['Total Classes']) * 100
    return df

# Function to consolidate uploaded attendance data
def consolidate_attendance_data(files):
    attendance_data = []

    for file in files:
        try:
            # Read the uploaded Excel file
            df = pd.read_excel(file)

            # Standardize column names
            df.columns = [col.strip().lower() for col in df.columns]

            # Rename columns to a standard format
            rename_map = {
                'student name': 'Name',
                'total classes': 'Total Classes',
                'classes attended': 'Classes Attended'
            }
            df.rename(columns=rename_map, inplace=True)

            # Check if required columns are present after renaming
            if 'Total Classes' in df.columns and 'Classes Attended' in df.columns:
                df = calculate_attendance_percentage(df)
                attendance_data.append(df)
            else:
                st.error(f"Required columns not found in the file {file.name}.")

        except Exception as e:
            st.error(f"Error reading file {file.name}: {e}")

    if not attendance_data:
        st.warning("No valid data to concatenate.")
        return pd.DataFrame()

    # Concatenate all dataframes into a single dataframe
    consolidated_df = pd.concat(attendance_data, ignore_index=True)
    return consolidated_df

# Function to highlight students with attendance below 75%
def highlight_low_attendance(df):
    highlight_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    df.style.applymap(lambda x: 'background-color: #FF9999' if x < 75 else '', subset=['Attendance Percentage'])
    return df

# Main function to create the Streamlit app
def main():
    st.title("Attendance Consolidation App")
    st.write("Upload Excel files to consolidate attendance data, calculate attendance percentages, and highlight students with low attendance.")

    uploaded_files = st.file_uploader("Upload Attendance Excel Files", accept_multiple_files=True, type=["xlsx"])

    if uploaded_files:
        consolidated_df = consolidate_attendance_data(uploaded_files)

        if not consolidated_df.empty:
            st.success("Consolidation Successful!")
            st.write("Consolidated Data Preview:")
            st.dataframe(consolidated_df)

            # Highlight students with attendance below 75%
            highlighted_df = highlight_low_attendance(consolidated_df)

            # Save the consolidated data to a downloadable Excel file
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                consolidated_df.to_excel(writer, index=False, sheet_name='Consolidated Report')
                workbook = writer.book
                worksheet = writer.sheets['Consolidated Report']

                # Apply conditional formatting for low attendance
                format_red = workbook.add_format({'bg_color': '#FF9999'})
                worksheet.conditional_format('F2:F{}'.format(len(consolidated_df) + 1), {'type': 'cell', 'criteria': '<', 'value': 75, 'format': format_red})

            st.download_button(
                label="Download Consolidated Report",
                data=output.getvalue(),
                file_name="consolidated_attendance_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
if __name__ == " cd__main__":
    main()
